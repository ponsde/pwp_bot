"""Pipeline entrypoint for ETL and answer tasks."""
from __future__ import annotations

import argparse
import json
import logging
from pathlib import Path

import openpyxl

from config import OV_CONFIG_PATH, OV_DATA_DIR, REPORTS_DIR, RESEARCH_QUESTIONS_XLSX

logger = logging.getLogger(__name__)


def _safe_chart_data(rows: list[dict]) -> list[dict]:
    """Build chart-friendly data, handling non-numeric first columns (e.g. report_period)."""
    data = []
    for row in rows:
        items = list(row.items())
        if len(items) >= 2:
            label = str(items[0][1])
            value = items[1][1]
        elif len(items) == 1:
            label = str(len(data) + 1)
            value = items[0][1]
        else:
            continue
        try:
            data.append({"label": label, "value": float(value)})
        except (ValueError, TypeError):
            continue
    return data


def run_etl(input_dir: str, db_path: str) -> dict[str, object]:
    from src.etl.loader import ETLLoader
    loader = ETLLoader(Path(db_path))
    pdf_paths = sorted(Path(input_dir).rglob("*.pdf"))
    results = []
    for pdf_path in pdf_paths:
        try:
            result = loader.load_pdf(pdf_path)
        except Exception as exc:
            result = {"status": "error", "file": str(pdf_path), "error": str(exc), "reason": str(exc)}
        else:
            if "reason" not in result:
                if result["status"] == "loaded":
                    result["reason"] = "ok"
                elif result["status"] == "skipped":
                    result["reason"] = result.get("reason", "skipped")
                elif result["status"] == "rejected":
                    warnings = result.get("warnings") or []
                    result["reason"] = "; ".join(warnings) if warnings else "validation_failed"
        results.append(result)
        logger.info("ETL %s | %s | %s", result.get("status"), result.get("file"), result.get("reason", ""))
    return {
        "status": "completed_with_errors" if any(r["status"] in {"error", "rejected"} for r in results) else "completed",
        "db_path": db_path,
        "input_dir": input_dir,
        "processed": len(results),
        "loaded": sum(1 for r in results if r["status"] == "loaded"),
        "skipped": sum(1 for r in results if r["status"] == "skipped"),
        "rejected": sum(1 for r in results if r["status"] == "rejected"),
        "error": sum(1 for r in results if r["status"] == "error"),
        "results": results,
    }


def _load_questions_xlsx(path: str) -> list[dict]:
    """Read questions xlsx (附件4 format): columns 编号, 问题类型, 问题."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    questions = []
    for row in rows:
        if not row or not row[0]:
            continue
        question_id = str(row[0]).strip()
        question_json = str(row[2]).strip() if row[2] else "[]"
        turns = json.loads(question_json)
        questions.append({"id": question_id, "turns": turns})
    return questions


def _load_research_questions_xlsx(path: str) -> list[dict]:
    """Read research questions xlsx (附件6 format): columns 编号, 问题类型, 问题."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    questions = []
    for row in rows:
        if not row or not row[0]:
            continue
        questions.append({
            "id": str(row[0]).strip(),
            "question_type": str(row[1]).strip() if row[1] else "",
            "turns": json.loads(str(row[2]).strip()) if row[2] else [],
        })
    return questions


def run_research(questions_path: str, db_path: str, output_xlsx: str) -> str:
    import pandas as pd

    from src.knowledge.ov_adapter import init_client
    from src.knowledge.research_loader import load_research_documents
    from src.knowledge.research_qa import ResearchQAEngine, format_research_answer_payload
    from src.query.chart import render_chart
    from src.query.conversation import ConversationManager

    client = init_client(data_path=OV_DATA_DIR, config_path=OV_CONFIG_PATH)
    load_research_documents(client)
    engine = ResearchQAEngine(db_path=db_path, client=client, llm_client=_build_llm_client())
    rows = []
    for item in _load_research_questions_xlsx(questions_path):
        sql_parts = []
        answer_payloads = []
        chart_type = "无"
        conversation = ConversationManager()
        for turn_index, turn in enumerate(item["turns"], start=1):
            chart_image: str | None = None
            question = turn["Q"]
            conversation.add_user_message(question)
            answer = engine.answer_question(question, conversation)
            if answer.sql:
                sql_parts.append(answer.sql)
            if answer.chart_type and answer.chart_type != "无":
                chart_type = answer.chart_type
                chart_rows = getattr(answer, "chart_rows", [])
                if chart_rows:
                    chart_data = _safe_chart_data(chart_rows)
                    chart_image = render_chart(
                        answer.chart_type,
                        chart_data,
                        f"result/{item['id']}_{turn_index}.jpg",
                        question,
                    ) if chart_data else None
            answer_payload = json.loads(format_research_answer_payload(answer))
            if chart_image:
                answer_payload.setdefault("image", [chart_image])
            answer_payloads.append(answer_payload)
            conversation.add_assistant_message(answer.answer)
        rows.append({
            "编号": item["id"],
            "问题": json.dumps(item["turns"], ensure_ascii=False),
            "SQL查询语句": "\n\n".join(sql_parts),
            "图形格式": chart_type,
            "回答": json.dumps(answer_payloads, ensure_ascii=False),
        })
    path = Path(output_xlsx)
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_excel(path, index=False)
    return str(path)

def _build_llm_client():
    from src.llm.client import LLMClient

    try:
        return LLMClient.from_env()
    except (ValueError, OSError, RuntimeError) as exc:
        logger.warning("LLM client unavailable, fallback to heuristic mode: %s", exc)
        return None


def run_answer(questions_path: str, db_path: str, output_xlsx: str) -> str:
    from src.query.answer import build_answer_content, build_answer_record
    from src.query.chart import render_chart, select_chart_type
    from src.query.conversation import ConversationManager
    from src.query.text2sql import Text2SQLEngine

    engine = Text2SQLEngine(db_path=db_path, llm_client=_build_llm_client())
    all_records: list[dict] = []

    questions = _load_questions_xlsx(questions_path)
    for q_item in questions:
        question_id = q_item["id"]
        turns = q_item["turns"]
        conversation = ConversationManager()
        turn_records: list[dict] = []

        for turn in turns:
            question = turn["Q"]
            conversation.add_user_message(question)
            try:
                result = engine.query(question, conversation)
            except Exception as exc:
                logger.exception("Failed to answer question %s turn %s", question_id, question)
                from src.query.text2sql import QueryResult
                result = QueryResult(sql=None, rows=[], intent={}, error=f"查询失败：{exc}")

            if result.needs_clarification:
                content = result.clarification_question or "请补充信息。"
                chart_type_str = "none"
                images: list[str] = []
            elif result.error:
                content = result.error
                chart_type_str = "none"
                images = []
            else:
                chart_data = _safe_chart_data(result.rows)
                chart_type_str = select_chart_type(question, chart_data)
                image = render_chart(
                    chart_type_str,
                    chart_data,
                    f"result/{question_id}_{len(turn_records) + 1}.jpg",
                    question,
                ) if chart_data else None
                images = [image] if image else []
                content = build_answer_content(question, result.rows)
                if result.warning:
                    content += f"\n（注：{result.warning}）"

            conversation.add_assistant_message(content)
            turn_records.append(
                build_answer_record(
                    question,
                    content,
                    images,
                    chart_type_str,
                    sql=result.sql or "",
                )
            )

        all_records.append({
            "id": question_id,
            "turns_input": turns,
            "turn_records": turn_records,
        })

    return _write_grouped_result_xlsx(all_records, output_xlsx)


def _write_grouped_result_xlsx(grouped: list[dict], output_path: str) -> str:
    """Write result_2.xlsx in 附件7 format: one row per question group."""
    import pandas as pd

    rows = []
    for item in grouped:
        question_id = item["id"]
        turns_input = item["turns_input"]
        turn_records = item["turn_records"]
        first_sql = ""
        chart_type_label = "无"
        for rec in turn_records:
            if rec.get("sql"):
                first_sql = rec["sql"]
            if rec.get("chart_type", "无") != "无":
                chart_type_label = rec["chart_type"]
        answer_json = json.dumps(
            [{"Q": rec["Q"], "A": rec["A"]} for rec in turn_records],
            ensure_ascii=False,
        )
        rows.append({
            "编号": question_id,
            "问题": json.dumps(turns_input, ensure_ascii=False),
            "SQL查询语句": first_sql,
            "图形格式": chart_type_label,
            "回答": answer_json,
        })
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_excel(path, index=False)
    return str(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Financial Report QA Pipeline")
    parser.add_argument("--task", choices=["etl", "answer", "research"], required=True)
    parser.add_argument("--db-path", default="data/db/finance.db")
    parser.add_argument("--input", default=str(REPORTS_DIR), help="PDF directory for ETL")
    parser.add_argument("--questions", default=None, help="Questions xlsx for answer task")
    parser.add_argument("--output", default="result_2.xlsx")
    args = parser.parse_args()

    Path(args.db_path).parent.mkdir(parents=True, exist_ok=True)
    if args.task == "etl":
        print(json.dumps(run_etl(args.input, args.db_path), ensure_ascii=False, indent=2))
    elif args.task == "answer":
        if not args.questions:
            raise SystemExit("--questions is required for answer task")
        print(run_answer(args.questions, args.db_path, args.output))
    else:
        questions = args.questions or str(RESEARCH_QUESTIONS_XLSX)
        print(run_research(questions, args.db_path, args.output))


if __name__ == "__main__":
    main()
