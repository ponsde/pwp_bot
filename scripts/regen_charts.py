"""Regenerate chart images for result_2.xlsx / result_3.xlsx using the new chart style.

For each row whose 回答 JSON contains an A.image entry, re-execute the associated
SQL (from 'SQL查询语句' or 'SQL查询语法' column) and render a fresh chart at the
same output path. No-op for rows without images/SQL.
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from src.query.chart import render_chart, safe_chart_data, select_chart_type

try:
    from src.llm.client import LLMClient
    _llm = LLMClient.from_env()
except Exception:
    _llm = None

_TITLE_CACHE: dict[str, str] = {}


def _llm_title(question: str) -> str:
    """Ask LLM for a short chart title (<= 20 chars). Heuristic fallback if LLM unavailable."""
    if question in _TITLE_CACHE:
        return _TITLE_CACHE[question]
    if _llm is not None:
        prompt = (
            "为下面这个财报数据可视化图表生成一个**简洁的中文标题**（10-18 字）。"
            "只返回纯标题，不要引号、不要解释、不要问号。\n\n"
            f"问题：{question.strip()}"
        )
        try:
            title = str(_llm.complete(prompt)).strip()
            title = title.strip('"“”\'「」《》。！？?!:：')
            if 2 <= len(title) <= 30:
                _TITLE_CACHE[question] = title
                return title
        except Exception:
            pass
    # Heuristic fallback — strip command tails
    t = question.strip()
    for tail in ("做可视化绘图", "请绘制", "请画图", "画图", "生成趋势折线图",
                 "生成折线图", "生成柱状图", "生成图表", "用图表展示",
                 "展示各报告期的数据", "是什么样的"):
        if tail in t:
            t = t.split(tail)[0].rstrip("，,。；;：:")
    return t[:24]


def _execute_sql(db: sqlite3.Connection, sql: str) -> list[dict]:
    try:
        cur = db.execute(sql)
        cols = [d[0] for d in cur.description] if cur.description else []
        return [dict(zip(cols, row)) for row in cur.fetchall()]
    except sqlite3.Error as exc:
        print(f"  SQL error: {exc}")
        return []


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--db", default="data/db/finance.db")
    parser.add_argument("--result-dir", default="result")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    sql_col = "SQL查询语句" if "SQL查询语句" in headers else "SQL查询语法"

    db = sqlite3.connect(args.db)
    regenerated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2):
        values = {h: c.value for h, c in zip(headers, row)}
        qid = str(values.get("编号") or "").strip()
        sql_text = str(values.get(sql_col) or "").strip()
        ans_text = str(values.get("回答") or "").strip()
        if not qid or not ans_text:
            continue
        try:
            ans_obj = json.loads(ans_text)
        except Exception:
            continue

        # Multiple SQLs may be joined by \n\n — take the first statement that
        # looks chart-worthy (multiple rows). Also try splitting by ';'.
        sql_parts = [s.strip() for s in re.split(r"\n\s*\n|;\s*\n", sql_text) if s.strip()]
        if not sql_parts:
            continue
        # Re-render each turn that has images
        for turn_idx, turn in enumerate(ans_obj, start=1):
            A = turn.get("A", {})
            imgs = A.get("image") or []
            if not imgs:
                continue
            # Pick a sql — use the first one that returns multi-row
            chosen_rows: list[dict] = []
            chosen_sql = ""
            for sql in sql_parts:
                rows = _execute_sql(db, sql)
                if len(rows) > len(chosen_rows):
                    chosen_rows = rows
                    chosen_sql = sql
            if not chosen_rows or len(chosen_rows) < 2:
                skipped += 1
                print(f"[{qid}_{turn_idx}] skip: insufficient rows ({len(chosen_rows)})")
                continue
            chart_data, value_field = safe_chart_data(chosen_rows)
            chart_type = select_chart_type(turn.get("Q", ""), chart_data)
            if not chart_data or chart_type == "none":
                skipped += 1
                print(f"[{qid}_{turn_idx}] skip: chart_type=none")
                continue
            target = Path(args.result_dir) / f"{qid}_{turn_idx}.jpg"
            title = _llm_title(turn.get("Q", ""))
            out = render_chart(
                chart_type,
                chart_data,
                str(target),
                title=title,
                value_field=value_field,
            )
            if out:
                regenerated += 1
                print(f"[{qid}_{turn_idx}] regenerated {chart_type} -> {target}")
            else:
                skipped += 1
    db.close()
    wb.close()
    print(f"\n=== DONE: regenerated={regenerated} skipped={skipped} ===")


if __name__ == "__main__":
    main()
