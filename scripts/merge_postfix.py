"""Overlay postfix-run answers onto a base result xlsx.

Reads the source xlsx (e.g. result/postfix_8/result_3.xlsx), finds every
row whose answer is non-empty AND doesn't start with "查询失败"/"处理失败",
and overwrites the matching ID row in the destination xlsx. Also copies
any matching chart jpgs from src_dir to dst_dir.

Usage:
    python3 scripts/merge_postfix.py <src_xlsx> <src_dir> <dst_xlsx> <dst_dir>
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl


def _is_ok_answer(cell_value) -> bool:
    if not cell_value:
        return False
    s = str(cell_value)
    if "查询失败" in s or "处理失败" in s:
        return False
    return True


def main(argv: list[str]) -> int:
    if len(argv) != 5:
        print(__doc__)
        return 2
    src_xlsx = Path(argv[1])
    src_dir = Path(argv[2])
    dst_xlsx = Path(argv[3])
    dst_dir = Path(argv[4])

    src_wb = openpyxl.load_workbook(src_xlsx)
    src_ws = src_wb.active
    replacements: dict[str, tuple] = {}
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        qid = str(row[0]).strip()
        if _is_ok_answer(row[3]):
            replacements[qid] = row
    src_wb.close()

    dst_wb = openpyxl.load_workbook(dst_xlsx)
    dst_ws = dst_wb.active
    touched = []
    for ws_row in dst_ws.iter_rows(min_row=2):
        qid_cell = ws_row[0]
        qid = str(qid_cell.value).strip() if qid_cell.value else ""
        if qid in replacements:
            new_row = replacements[qid]
            for idx, cell in enumerate(ws_row):
                if idx < len(new_row):
                    cell.value = new_row[idx]
            touched.append(qid)

    dst_wb.save(dst_xlsx)
    dst_wb.close()

    copied = []
    for qid in touched:
        for src_jpg in src_dir.glob(f"{qid}_*.jpg"):
            target = dst_dir / src_jpg.name
            shutil.copy2(src_jpg, target)
            copied.append(src_jpg.name)

    print(f"replaced rows: {touched}")
    print(f"copied charts: {copied}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
