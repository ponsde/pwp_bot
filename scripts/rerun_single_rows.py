"""Re-run specific question IDs through vikingbot and patch them into an
existing result xlsx in place.

Usage:
  .venv/bin/python scripts/rerun_single_rows.py \
      --task research \
      --questions "附件6：问题汇总.xlsx" \
      --result-xlsx result_3.xlsx \
      --only B2019,B2026,B2029,B2046
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
import requests  # noqa: E402

# Reuse the bot machinery from batch_via_bot
from scripts.batch_via_bot import (  # noqa: E402
    DEFAULT_API_KEY,
    DEFAULT_GATEWAY,
    _bot_run_question,
    _load_questions_xlsx,
)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--task", choices=["answer", "research"], required=True)
    ap.add_argument("--questions", required=True, help="Path to 附件4 / 附件6 xlsx")
    ap.add_argument("--result-xlsx", required=True)
    ap.add_argument("--only", required=True, help="Comma-separated B-ids")
    ap.add_argument("--gateway", default=DEFAULT_GATEWAY)
    ap.add_argument("--api-key", default=DEFAULT_API_KEY)
    ap.add_argument("--result-dir", default="result")
    args = ap.parse_args()

    only = {s.strip() for s in args.only.split(",") if s.strip()}
    if not only:
        print("no --only ids supplied", file=sys.stderr)
        return 1

    try:
        requests.get(f"{args.gateway}/bot/v1/health", timeout=10).raise_for_status()
    except Exception as exc:
        print(f"gateway unreachable: {exc}", file=sys.stderr)
        return 2

    items = [it for it in _load_questions_xlsx(Path(args.questions)) if it["id"] in only]
    if not items:
        print("no matching questions found in input xlsx", file=sys.stderr)
        return 3

    print(f"rerunning {len(items)} questions: {[i['id'] for i in items]}")
    result_dir = Path(args.result_dir)
    result_dir.mkdir(parents=True, exist_ok=True)

    reruns: dict[str, dict] = {}
    for i, item in enumerate(items, 1):
        t0 = time.time()
        try:
            r = _bot_run_question(args.gateway, args.api_key, item, result_dir, args.task)
        except Exception as exc:
            print(f"[{item['id']}] FAIL: {exc}", file=sys.stderr)
            continue
        reruns[item["id"]] = r
        print(f"[{i}/{len(items)}] {item['id']} ok in {time.time()-t0:.1f}s")

    # Patch back into the result xlsx
    xlsx_path = Path(args.result_xlsx)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    bh_col = headers.index("编号") + 1
    q_col = headers.index("问题") + 1
    sql_col_name = "SQL查询语句" if "SQL查询语句" in headers else "SQL查询语法"
    sql_col = headers.index(sql_col_name) + 1
    ans_col = headers.index("回答") + 1
    chart_col = headers.index("图形格式") + 1 if "图形格式" in headers else None

    patched = 0
    for r in range(2, ws.max_row + 1):
        bh = ws.cell(r, bh_col).value
        if bh not in reruns:
            continue
        data = reruns[bh]
        ws.cell(r, q_col).value = json.dumps(data["turns_input"], ensure_ascii=False)
        ws.cell(r, sql_col).value = data["sql_joined"]
        ws.cell(r, ans_col).value = json.dumps(data["answer_payloads"], ensure_ascii=False)
        if chart_col is not None:
            ws.cell(r, chart_col).value = data["chart_type"]
        patched += 1

    out = xlsx_path.with_suffix(".patched.xlsx")
    wb.save(out)
    print(f"wrote {out}  (patched {patched} rows; review then mv over {xlsx_path.name})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
