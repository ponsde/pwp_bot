"""Audit result xlsx files and emit paper/audit_report.md.

Re-runs every row's SQL against the DB, extracts numbers from the content,
checks chart files, validates references, and — for task-3 Hybrid/multi-intent
rows — runs an LLM-as-judge narrative score. Produces a tri-severity
Markdown summary at paper/audit_report.md by default.
"""
from __future__ import annotations

import argparse
import ast
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from src.audit.checks import Finding, check_chart_file, check_number_consistency  # noqa: E402
from src.audit.llm_judge import judge_narrative  # noqa: E402
from src.audit.reference_validator import validate_reference  # noqa: E402
from src.audit.report import render_report  # noqa: E402
from src.audit.sql_runner import SqlRunError, run_sql_strict  # noqa: E402


def _extract_answers(raw: str) -> list[dict]:
    if not raw:
        return []
    try:
        obj = json.loads(raw)
    except ValueError:
        try:
            obj = ast.literal_eval(raw)
        except (ValueError, SyntaxError):
            return []
    return obj if isinstance(obj, list) else []


_IMG_RE = re.compile(r"\./result/(B\d+_\d+\.jpg)")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--db", default=str(ROOT / "data" / "db" / "finance.db"))
    ap.add_argument("--out", default=str(ROOT / "paper" / "audit_report.md"))
    ap.add_argument("--no-judge", action="store_true", help="skip LLM-as-judge")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    db_path = Path(args.db)
    out_path = Path(args.out)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    sql_col_name = "SQL查询语句" if "SQL查询语句" in headers else "SQL查询语法"
    idx = {h: i for i, h in enumerate(headers)}

    llm = None
    if not args.no_judge:
        try:
            from src.llm.client import LLMClient
            llm = LLMClient.from_env()
        except Exception:
            llm = None

    findings: list[Finding] = []
    totals = {"blocking": 0, "suspect": 0, "hint": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        bh = str(row[idx["编号"]] or "").strip()
        if not bh:
            continue
        sql = str(row[idx.get(sql_col_name, 2)] or "").strip()
        if sql in {"无", "无。"}:
            sql = ""
        answer_raw = str(row[idx.get("回答", len(headers) - 1)] or "")

        try:
            sql_rows = run_sql_strict(db_path, sql) if sql else []
        except SqlRunError as exc:
            findings.append(
                Finding(bh, "blocking", "sql_error", f"SQL execution failed: {exc}")
            )
            sql_rows = []

        answers = _extract_answers(answer_raw)
        for turn in answers:
            a = (turn or {}).get("A") or {}
            content = str(a.get("content") or "")
            for f in check_number_consistency(bh, content=content, sql_rows=sql_rows):
                findings.append(f)
            for m in _IMG_RE.finditer(content):
                img_path = ROOT / "result" / m.group(1)
                for f in check_chart_file(bh, path=img_path):
                    findings.append(f)
            for img in (a.get("image") or []):
                img_name = Path(str(img)).name
                img_path = ROOT / "result" / img_name
                for f in check_chart_file(bh, path=img_path):
                    findings.append(f)
            refs = a.get("references") or []
            refs_text: list[str] = []
            for ref in refs:
                if not isinstance(ref, dict):
                    continue
                refs_text.append(str(ref.get("text") or ""))
                rr = validate_reference(
                    ref=ref,
                    repo_root=ROOT,
                    ov_root=ROOT / ".openviking",
                )
                if not rr.path_ok:
                    findings.append(
                        Finding(
                            bh, "suspect", "ref_path_missing",
                            f"paper_path not found: {ref.get('paper_path')!r}",
                        )
                    )
                elif not rr.text_ok:
                    findings.append(
                        Finding(
                            bh, "suspect", "ref_text_miss",
                            "reference text not found in any OV chunk",
                        )
                    )
            if content and len(content.strip()) < 30:
                findings.append(
                    Finding(bh, "hint", "short_content", f"content length {len(content)}")
                )
            if llm is not None and refs and not args.no_judge:
                v = judge_narrative(
                    question=str(turn.get("Q") or ""),
                    content=content,
                    references_text=refs_text,
                    llm=llm,
                )
                if v.is_weak():
                    findings.append(
                        Finding(bh, "blocking", "narrative_too_weak",
                                f"LLM judged {v.score}/3: {v.reason}")
                    )

    for f in findings:
        totals[f.severity] += 1
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(render_report(findings, totals=totals), encoding="utf-8")
    print(f"wrote {out_path}")
    print(f"  {totals}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
