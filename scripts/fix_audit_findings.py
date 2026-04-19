"""Auto-apply fixes based on paper/audit_report*.md findings.

Subcommands:
  clean-refs    Drop references with empty / broken paper_path
  rewrite       Ask LLM to rewrite content using SQL rows as ground truth

Writes updated xlsx to <input>.audited.xlsx so the original is preserved.
"""
from __future__ import annotations

import argparse
import ast
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from src.audit.reference_validator import validate_reference  # noqa: E402


def _load_cell(raw: str):
    if not raw:
        return None
    try:
        return json.loads(raw)
    except ValueError:
        try:
            return ast.literal_eval(raw)
        except (ValueError, SyntaxError):
            return None


def _dump_cell(obj) -> str:
    return json.dumps(obj, ensure_ascii=False)


def clean_broken_refs_in_cell(raw: str, *, repo_root: Path, ov_root: Path) -> str:
    obj = _load_cell(raw)
    if not isinstance(obj, list):
        return raw
    for turn in obj:
        if not isinstance(turn, dict):
            continue
        A = turn.get("A") or {}
        refs = A.get("references") or []
        kept = []
        for ref in refs:
            if not isinstance(ref, dict):
                continue
            pp = str(ref.get("paper_path") or "").strip()
            if not pp or pp.startswith("viking://"):
                continue  # drop empties and internal viking URIs
            rr = validate_reference(ref=ref, repo_root=repo_root, ov_root=ov_root)
            if rr.path_ok:
                kept.append(ref)
        A["references"] = kept
        turn["A"] = A
    return _dump_cell(obj)


def _clean_refs(xlsx_in: Path, xlsx_out: Path, repo_root: Path, ov_root: Path) -> int:
    wb = openpyxl.load_workbook(xlsx_in)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ans_col = headers.index("回答") + 1
    changed = 0
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, ans_col)
        before = str(cell.value or "")
        after = clean_broken_refs_in_cell(before, repo_root=repo_root, ov_root=ov_root)
        if after != before:
            cell.value = after
            changed += 1
    wb.save(xlsx_out)
    return changed


_REWRITE_PROMPT = (
    "你是财报问答助手。基于下列信息重写 content（≤200 字），"
    "必须原样引用 SQL 行的数字（保留小数），不得编造其他数字：\n"
    "问题：{q}\n"
    "SQL 返回行：{s}\n"
    "原 content：{c}\n"
    "只返回重写后的 content 文本，不加引号，不解释。"
)


def rewrite_content_in_cell(raw: str, *, sql_rows: list[dict], llm, weak_ids: set[int]) -> str:
    obj = _load_cell(raw)
    if not isinstance(obj, list):
        return raw
    for i, turn in enumerate(obj):
        if i not in weak_ids or not isinstance(turn, dict):
            continue
        A = turn.get("A") or {}
        prompt = _REWRITE_PROMPT.format(
            q=str(turn.get("Q") or ""),
            s=json.dumps(sql_rows[:10], ensure_ascii=False),
            c=str(A.get("content") or ""),
        )
        try:
            new_content = str(llm.complete(prompt)).strip()
        except Exception:
            continue
        if new_content:
            A["content"] = new_content
            turn["A"] = A
    return _dump_cell(obj)


def _parse_report_for_row_ids(report_path: Path, kinds: set[str]) -> dict[str, set[int]]:
    if not report_path.exists():
        return {}
    out: dict[str, set[int]] = {}
    for line in report_path.read_text(encoding="utf-8").splitlines():
        if not line.startswith("| "):
            continue
        parts = [p.strip() for p in line.strip("|").split("|")]
        if len(parts) < 3:
            continue
        bh, kind = parts[0], parts[1]
        if kind in kinds and bh.startswith("B"):
            out.setdefault(bh, set()).add(0)
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    p1 = sub.add_parser("clean-refs")
    p1.add_argument("--xlsx", required=True)
    p1.add_argument("--out")

    p2 = sub.add_parser("rewrite")
    p2.add_argument("--xlsx", required=True)
    p2.add_argument("--db", default=str(ROOT / "data" / "db" / "finance.db"))
    p2.add_argument("--report", required=True)
    p2.add_argument("--out")

    args = ap.parse_args()

    if args.cmd == "clean-refs":
        in_path = Path(args.xlsx)
        out_path = Path(args.out) if args.out else in_path.with_suffix(".audited.xlsx")
        n = _clean_refs(in_path, out_path, repo_root=ROOT, ov_root=ROOT / ".openviking")
        print(f"wrote {out_path}  (modified {n} rows)")
        return 0

    if args.cmd == "rewrite":
        from src.audit.sql_runner import run_sql_strict
        try:
            from src.llm.client import LLMClient
            llm = LLMClient.from_env()
        except Exception:
            print("LLM unavailable; aborting rewrite.", file=sys.stderr)
            return 2
        in_path = Path(args.xlsx)
        out_path = Path(args.out) if args.out else in_path.with_suffix(".rewritten.xlsx")
        weak = _parse_report_for_row_ids(
            Path(args.report), kinds={"num_mismatch", "narrative_too_weak"}
        )
        wb = openpyxl.load_workbook(in_path)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        sql_col_name = "SQL查询语句" if "SQL查询语句" in headers else "SQL查询语法"
        ans_col = headers.index("回答") + 1
        bh_col = headers.index("编号") + 1
        sql_col = headers.index(sql_col_name) + 1
        n = 0
        for r in range(2, ws.max_row + 1):
            bh = str(ws.cell(r, bh_col).value or "").strip()
            if bh not in weak:
                continue
            sql = str(ws.cell(r, sql_col).value or "").strip()
            try:
                sql_rows = run_sql_strict(args.db, sql) if sql and sql != "无" else []
            except Exception:
                sql_rows = []
            cell = ws.cell(r, ans_col)
            cell.value = rewrite_content_in_cell(
                str(cell.value or ""),
                sql_rows=sql_rows,
                llm=llm,
                weak_ids=weak[bh],
            )
            n += 1
        wb.save(out_path)
        print(f"wrote {out_path}  (rewrote {n} rows)")
        return 0
    return 1


if __name__ == "__main__":
    sys.exit(main())
