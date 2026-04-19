"""Rewrite paper_path values in result_3.xlsx to match 附件7 表6 format.

Bot returns viking uris like:
    ./resources/医药健康行业研究从2025医保谈判看行业风向成功率提升创新导向持续强化/事件简介/....md

Maps back to the real file path:
    ./附件5：研报数据/行业研报/医药健康行业研究：从 2025 医保谈判看行业风向：成功率提升，创新导向持续强化.pdf

Strategy: build an index from the PDF directory (sanitized-stem → original path),
then for every reference entry, look up the top-level dir from the viking uri,
strip the remainder (subsection/chunk .md), and map back.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RESEARCH_ROOT = ROOT / "data" / "sample" / "示例数据" / "附件5：研报数据"
_SANITIZE_STRICT = re.compile(r"[^\w\u4e00-\u9fff]")
# OV uses a softer sanitizer that keeps ASCII word chars AND hyphens
_SANITIZE_SOFT = re.compile(r"[^\w\u4e00-\u9fff\-]")
# OV also sometimes replaces whitespace with underscore (and strips other punctuation)
_WHITESPACE_RE = re.compile(r"\s+")
_PUNCT_RE = re.compile(r"[^\w\u4e00-\u9fff\-_]")


def _sanitize_strict(title: str) -> str:
    return _SANITIZE_STRICT.sub("", title)


def _sanitize_soft(title: str) -> str:
    return _SANITIZE_SOFT.sub("", title)


def _sanitize_underscore(title: str) -> str:
    # Whitespace → '_', other punctuation stripped, keep hyphens
    tmp = _WHITESPACE_RE.sub("_", title)
    tmp = _PUNCT_RE.sub("", tmp)
    return tmp


def _build_pdf_index() -> dict[str, str]:
    """Return {sanitized_stem: relative_path_under_root}.

    Builds keys under STRICT + SOFT + UNDERSCORE sanitizers so we catch OV's naming.
    """
    idx: dict[str, str] = {}
    for sub in ("个股研报", "行业研报"):
        for pdf in sorted((RESEARCH_ROOT / sub).glob("*.pdf")):
            rel = f"./附件5：研报数据/{sub}/{pdf.name}"
            for key in (_sanitize_strict(pdf.stem),
                        _sanitize_soft(pdf.stem),
                        _sanitize_underscore(pdf.stem)):
                if key and key not in idx:
                    idx[key] = rel
    return idx


def _extract_top_stem(raw_path: str) -> str | None:
    """From './resources/<sanitized>/sub/.../x.md' extract '<sanitized>'."""
    if not raw_path:
        return None
    # Strip leading ./ or viking://
    cleaned = raw_path
    for prefix in ("./", "viking://"):
        if cleaned.startswith(prefix):
            cleaned = cleaned[len(prefix):]
    # Strip leading "resources/" if present
    if cleaned.startswith("resources/"):
        cleaned = cleaned[len("resources/"):]
    parts = cleaned.split("/")
    return parts[0] if parts else None


def _resolve_path(raw_path: str, idx: dict[str, str]) -> str:
    """Map raw reference path to 附件5 format, fall back to raw if unknown."""
    stem = _extract_top_stem(raw_path)
    if not stem:
        return raw_path
    # Try the stem directly; some OV titles end with `_1 _2` suffix due to collision,
    # or `_<hexhash>` when sanitized name was too long.
    candidates = [stem]
    candidates.append(re.sub(r"_\d+$", "", stem))
    candidates.append(re.sub(r"_[a-f0-9]{6,}$", "", stem))
    # Both in sequence: strip trailing digits, then trailing hex
    stripped = re.sub(r"_\d+$", "", stem)
    stripped = re.sub(r"_[a-f0-9]{6,}$", "", stripped)
    candidates.append(stripped)
    for c in candidates:
        if c and c in idx:
            return idx[c]
    # Fuzzy: OV sometimes truncates. Find the longest key in idx that is a
    # prefix of `stem` (allowing OV to have added a suffix) or vice versa.
    stripped_base = stripped  # use the cleaned form
    if stripped_base:
        best_match = None
        best_len = 0
        for k, v in idx.items():
            if len(k) < 8:
                continue
            # k is the full sanitized original title; OV may have truncated it
            if stripped_base.startswith(k[:min(len(k), 40)]) or k.startswith(stripped_base[:40]):
                if len(k) > best_len:
                    best_match, best_len = v, len(k)
        if best_match:
            return best_match
    return raw_path


def fix_xlsx(path: Path, dry_run: bool = False) -> None:
    idx = _build_pdf_index()
    print(f"[idx] loaded {len(idx)} PDFs from {RESEARCH_ROOT}")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    answer_col = headers.index("回答") + 1

    total_refs = 0
    mapped = 0
    dropped_user_memory = 0
    unmapped_samples: list[str] = []

    for row in ws.iter_rows(min_row=2):
        ans_cell = row[answer_col - 1]
        try:
            payload = json.loads(ans_cell.value or "[]")
        except Exception:
            continue
        if not isinstance(payload, list):
            continue
        changed = False
        for turn in payload:
            A = turn.get("A") if isinstance(turn, dict) else None
            if not isinstance(A, dict):
                continue
            refs = A.get("references") or []
            new_refs = []
            for ref in refs:
                if not isinstance(ref, dict):
                    new_refs.append(ref)
                    continue
                total_refs += 1
                old_path = str(ref.get("paper_path") or "")
                # Drop user/session OV refs — they're internal state, not 研报 citations.
                if "viking://user/" in old_path or "viking://session/" in old_path or old_path.startswith("viking://skill"):
                    dropped_user_memory += 1
                    changed = True
                    continue
                new_path = _resolve_path(old_path, idx)
                if new_path != old_path:
                    mapped += 1
                    ref["paper_path"] = new_path
                    changed = True
                elif old_path and len(unmapped_samples) < 5:
                    unmapped_samples.append(old_path)
                new_refs.append(ref)
            A["references"] = new_refs
        if changed:
            ans_cell.value = json.dumps(payload, ensure_ascii=False)
    print(f"[stats] total_refs={total_refs} mapped={mapped} dropped_user_memory={dropped_user_memory}")

    if dry_run:
        print(f"[dry-run] would map {mapped}/{total_refs} references")
    else:
        wb.save(path)
        print(f"[saved] {path} | mapped {mapped}/{total_refs} references")

    if unmapped_samples:
        print("unmapped sample paths:")
        for s in unmapped_samples:
            print(" ", s[:200])


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="result_3.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    fix_xlsx(Path(args.xlsx), dry_run=args.dry_run)


if __name__ == "__main__":
    main()
